import openpyxl
from pathlib import Path
import matplotlib.pyplot as plt
import textwrap

# Pfad zur Excel-Datei (liegt im selben Ordner wie dieses Script)
CURRENT_DIR = Path(__file__).resolve().parent
INPUT_FILE = CURRENT_DIR / "Tabelle_Kategorienzuweisungen.xlsx"

# Reihenfolge der dargestellten Datensätze (Zeilen unterhalb IPCC)
SYSTEMS_ORDER = ["AS", "AV", "CORINE", "ESA"]

# IPCC-Kategorien, die entfernt werden sollen
EXCLUDE_IPCC = {"Settlements"}

# Layout-Parameter
MAX_CHARS_PER_LINE = 18
BASE_FONT_SIZE = 8
MIN_SEG_WIDTH_INCH = 1.3
WRAP_MIN_N = 10

IPCC_COLORS = {
    "Forestland":  "#228B22",
    "Cropland":    "#8b4513",
    "Grassland":   "#bcff1e",
    "Wetlands":    "#1e90ff",
    "Settlements": "#a9a9a9",
    "Otherlands":  "#f31383",
}

# Excel einlesen
wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
ws = wb["Tabelle1"] if "Tabelle1" in wb.sheetnames else wb.active

segments = {}
ipcc_order = []

current_ipcc = None
max_row = ws.max_row
max_col = ws.max_column

for r in range(1, max_row + 1):
    first = ws.cell(r, 1).value

    if first == "IPCC":
        current_ipcc = ws.cell(r, 2).value
        if current_ipcc not in ipcc_order:
            ipcc_order.append(current_ipcc)

    elif first in SYSTEMS_ORDER and current_ipcc:
        system = first
        subcats = []
        for c in range(2, max_col + 1):
            val = ws.cell(r, c).value
            if val is not None and str(val).strip() != "":
                subcats.append(str(val))
        if subcats:
            segments[(current_ipcc, system)] = subcats

# Excluded-IPCC entfernen
ipcc_order = [x for x in ipcc_order if x not in EXCLUDE_IPCC]

# global_max_n bleibt korrekt
global_max_n = max(len(v) for v in segments.values())

def wrap_label(text, max_chars):
    lines = textwrap.wrap(text, max_chars, break_long_words=False)
    return "\n".join(lines[:2])

def font_size_for_n(n):
    if n <= 15:
        return BASE_FONT_SIZE
    elif n <= 30:
        return BASE_FONT_SIZE - 1
    return BASE_FONT_SIZE - 2

# Zeilenaufbau
rows = []

for ipcc in ipcc_order:
    rows.append((ipcc, "IPCC", [ipcc]))

    for sys in SYSTEMS_ORDER:
        key = (ipcc, sys)
        if key in segments:
            rows.append((ipcc, sys, segments[key]))

    if ipcc != ipcc_order[-1]:
        rows.append((ipcc, None, None))

num_rows = len(rows)

# Figurgröße dynamisch abhängig von excluded IPCC
fig_width = MIN_SEG_WIDTH_INCH * global_max_n
fig_height = 0.5 * num_rows   # automatisch kleiner wenn excluded

fig, ax = plt.subplots(figsize=(fig_width, fig_height))

ax.set_xlim(0, global_max_n)
ax.set_ylim(-0.5, num_rows - 0.5)
ax.set_xticks([])
ax.set_xlabel("")

ALL_SYSTEMS = ["IPCC"] + SYSTEMS_ORDER

yticks = []
yticklabels = []

for idx, (ipcc, system, subcats) in enumerate(rows):
    if system in ALL_SYSTEMS:
        yticks.append(idx)
        yticklabels.append(system)

ax.set_yticks(yticks)
ax.set_yticklabels(yticklabels, fontsize=BASE_FONT_SIZE)


ax.set_title("IPCC-Kategorien und Zuordnungen aus den anderen LULC-Datensätzen", fontsize=BASE_FONT_SIZE + 9, fontweight="bold")

# Zeichnen
for y, (ipcc, system, subcats) in enumerate(rows):
    if system is None or subcats is None:
        continue

    if system == "IPCC":
        color = IPCC_COLORS.get(ipcc, "lightgrey")
        ax.barh(y, global_max_n, left=0, color=color)
        ax.text(global_max_n / 2, y, subcats[0],
                ha="center", va="center",
                fontsize=BASE_FONT_SIZE + 2, fontweight="bold")
        continue

    n = len(subcats)
    unit = global_max_n / n
    left = 0
    fsize = font_size_for_n(n)
    color = IPCC_COLORS.get(ipcc, "lightgrey")

    for sub in subcats:
        ax.barh(y, unit, left=left, color=color,
                edgecolor="white", linewidth=1)

        if n >= WRAP_MIN_N and len(sub) > MAX_CHARS_PER_LINE:
            label = wrap_label(sub, MAX_CHARS_PER_LINE)
        else:
            label = sub

        ax.text(left + unit / 2, y, label,
                ha="center", va="center", fontsize=fsize)
        left += unit

ax.invert_yaxis()

plt.tight_layout()
plt.savefig("ipcc_mapping.png", dpi=300, bbox_inches="tight")
plt.close()